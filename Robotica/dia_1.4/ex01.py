#automação de planilhas

'''
Criar:
1 - Criar o DataFrame com os dados
2 - Calcular a média e criar uma nova coluna usando o Pandas
3 - Salvar em excel usando o Pandas
4 - Abrir o arquivo com o openpyxl para formatar
5 - Formatar o cabeçalho: negrito e com preenchimento cinza claro 
6 - Formatar as colunas númericas para 2 casas decimais
7 - Destacar notas abaixo de 7 com fundo vermelho
8 - Salvar as alterações 
'''

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# 1 - Criar o DataFrame com os dados
dados = {
    'Nome': ['Victor', 'Gabriela', 'Cecilia', 'Yara'],
    'Nota 1': [9.0, 10.0, 7.0, 8.0],
    'Nota 2': [6.0, 5.0, 10.0, 4.0]
}
df = pd.DataFrame(dados)

# 2 - Calcular a média e criar uma nova coluna usando o Pandas
df['Média'] = df[['Nota 1', 'Nota 2']].mean(axis=1)


# 3 - Salvar em excel usando o Pandas
arquivo_excel = 'notas.xlsx'
df.to_excel(arquivo_excel, index=False, sheet_name='Notas')


# 4 - Abrir o arquivo com o openpyxl para formatar
wb = load_workbook(arquivo_excel)
ws = wb['Notas']


# 5 - Formatar o cabeçalho: negrito e com preenchimento cinza claro
header_fill = PatternFill(start_color='DDDDDD', fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill


# 6 - Formatar as colunas númericas para 2 casas decimais
#(row = linha)
for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
    for cell in row:
        cell.number_format = '0.00'


# 7 - Destacar notas abaixo de 7 com fundo vermelho
for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
    for cell in row:
        if isinstance(cell.value, (int, float)) and cell.value < 7:
            cell.fill = PatternFill(start_color='FFCCCC', fill_type='solid')


# 8 - Salvar as alterações
wb.save(arquivo_excel)
print(f'A planilha {arquivo_excel} foi criada e formatada com sucesso!')


