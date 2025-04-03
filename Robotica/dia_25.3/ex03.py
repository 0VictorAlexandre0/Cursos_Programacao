#Manipulação de arquivos(CSV e Excel)
#Criando arquivos CSV

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

#criando um DataFrame
livros = [
    {'Título': 'Python para iniciantes', 'Autor': 'Victor', 'Ano': 2025},
    {'Título': 'Automação com Python', 'Autor': 'Gabriela', 'Ano': 2022},
    {'Título': 'Python para Análise de dados', 'Autor': 'Cecília Martuzzo Rodrigues', 'Ano': 2035}
]

df = pd.DataFrame(livros)
print("DataFrame criado:")
print(df)


#exportar para o CSV
df.to_csv("livros.csv", index=False)
print('Arquivo CSV criado com sucesso!')


#exportar para Excel
df.to_excel("livros.xlsx", index=False)
print('Arquivo Excel criado com sucesso!')


#Ler o arquivo Excel e manipular os dados
df_novo = pd.read_excel('livros.xlsx')
df_novo['Ano'] = df_novo['Ano'] + 1
df_novo.to_excel('livros_atualizados.xlsx', index=False)
print('Arquivo atualizado com sucesso!')


#Formatar o arquivo Excel
wb = load_workbook('livros_atualizados.xlsx')
ws = wb.active #ws = worksheet

#Estilo para os Títulos
header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for celula in ws[1]:
    celula.fill = header_fill
    celula.font = header_font

wb.save('livros_formatados.xlsx')
print('Planilha formatada salva!')



